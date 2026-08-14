import io
from typing import List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel_report(columns: List[str], rows: List[List[Any]], title: str = "Analytix AI Hisoboti") -> io.BytesIO:
    """Generates a styled Excel workbook in memory for direct HTTP streaming download."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analytix AI Report"

    # Styles
    title_font = Font(name="Calibri", size=14, bold=True, color="1E2761")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E2761", end_color="1E2761", fill_type="solid")
    data_font = Font(name="Calibri", size=11, color="12163C")
    thin_border = Border(
        left=Side(style='thin', color='E2E6EC'),
        right=Side(style='thin', color='E2E6EC'),
        top=Side(style='thin', color='E2E6EC'),
        bottom=Side(style='thin', color='E2E6EC')
    )

    # Title row
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.row_dimensions[1].height = 25

    # Headers
    header_row_idx = 3
    ws.row_dimensions[header_row_idx].height = 24
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data Rows
    current_row = 4
    for row_data in rows:
        ws.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
